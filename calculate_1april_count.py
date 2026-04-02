import pandas as pd
import os
import re

def get_historical_data(file_paths):
    emails = set()
    usernames = set()
    for f in file_paths:
        if os.path.exists(f):
            try:
                df = pd.read_excel(f)
                email_col = None
                user_col = None
                for col in df.columns:
                    col_lower = str(col).lower()
                    if "email" in col_lower: email_col = col
                    if any(p in col_lower for p in ["username", "name", "profile"]): user_col = col
                
                if email_col:
                    emails.update(df[email_col].dropna().astype(str).str.strip().tolist())
                if user_col:
                    usernames.update(df[user_col].dropna().astype(str).str.strip().tolist())
            except Exception as e:
                # print(f"Error reading {f}: {e}")
                pass
    return emails, usernames

def is_junk_email(email):
    if not email: return True
    email = str(email).lower().strip()
    if '@' not in email or '.' not in email or '?' in email: return True
    if any(email.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp']): return True
    if any(k in email for k in ["sentry", "wixpress", "@2x", "placeholder", "example.com"]): return True
    generic_prefixes = ["sales", "info", "support", "admin", "contact", "enquiries", "jobs", "press", "hello", "marketing", "business", "help", "enquiry"]
    local_part = email.split('@')[0]
    if any(local_part == p or local_part.startswith(p + ".") or local_part.startswith(p + "@") for p in generic_prefixes): return True
    if len(local_part) > 20:
        if local_part.count('-') >= 3: return True
        hex_chars = re.findall(r'[0-9a-f]', local_part)
        if len(hex_chars) / len(local_part) > 0.7 and len(local_part) > 15: return True
    return False

def is_junk_username(username):
    if not username: return True
    username = str(username).strip()
    if re.search(r'\s\d+$', username): return True
    if len(username) > 30 and re.search(r'[0-9a-f]{10,}', username.lower()): return True
    return False

def clean_and_extract_email(email_str):
    if pd.isna(email_str): return None
    parts = re.split(r'[,\s;]+', str(email_str))
    for email in parts:
        email = email.strip()
        if is_junk_email(email): continue
        return email
    return None

# Files to check against
historical_files = [
    'final_creator_contacts.xlsx',
    'ignored_contacts.xlsx',
    '31March emails.xlsx',
    '30March emails.xlsx',
    '29March emails.xlsx',
    '28March emails.xlsx',
    '27March emails.xlsx',
    '26March emails.xlsx',
    '25March emails.xlsx',
    '24March emails.xlsx',
    '22March emails.xlsx',
    '21March emails.xlsx',
    '20March emails.xlsx',
    '19March emails.xlsx',
    '18March emails.xlsx',
    '17March emails.xlsx',
    '16th March emails.xlsx'
]
print(f"Loading historical data from {len(historical_files)} files...")
hist_emails, hist_usernames = get_historical_data(historical_files)

# New data fetched today
print("Loading new data from new_contacts.xlsx...")
if not os.path.exists('new_contacts.xlsx'):
    print("No new_contacts.xlsx found! Fetch data first.")
    exit(0)
    
df_new = pd.read_excel('new_contacts.xlsx')

filtered_rows = []
for _, row in df_new.iterrows():
    platform = row.get('Platform', '')
    username_raw = row.get('Username', '')
    email_raw = row.get('Email', '')
    
    email = clean_and_extract_email(email_raw)
    if is_junk_username(username_raw): continue
    username = str(username_raw).strip()
    if not email: continue
    
    if email in hist_emails or username in hist_usernames:
        continue
    
    filtered_rows.append({
        "Platform": platform,
        "Username": username,
        "Email": email
    })

df_final = pd.DataFrame(filtered_rows)
if not df_final.empty:
    df_final = df_final.drop_duplicates(subset=['Username'], keep='first')
    df_final = df_final.drop_duplicates(subset=['Email'], keep='first')

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def save_with_styling(df, path):
    if df.empty:
        df.to_excel(path, index=False)
        return

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12, name='Calibri')
        data_font = Font(size=11, name='Calibri')
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for i, col in enumerate(df.columns):
            # Column letter logic for more than 26 columns if needed
            column_letter = ""
            temp_i = i
            while temp_i >= 0:
                column_letter = chr(65 + (temp_i % 26)) + column_letter
                temp_i = (temp_i // 26) - 1
            
            max_length = max(len(str(col)) + 4, 15)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=i+1)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

if not df_final.empty:
    print(f"\nTotal truly fresh records for April 1st: {len(df_final)}")
else:
    print("\nTotal truly fresh records for April 1st: 0")

OUTPUT_FILE = "1April emails.xlsx"
save_with_styling(df_final, OUTPUT_FILE)
print(f"SUCCESS! Saved to {OUTPUT_FILE}")
