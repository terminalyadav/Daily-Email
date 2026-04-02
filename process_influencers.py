import pandas as pd
import os
import re
import sys
import gspread
from google.oauth2.service_account import Credentials

# Configuration
SPREADSHEET_ID = "1Q8Pw_a88RRZZ9dpaREMAgf5b0RqRqSkQdaJsLGWivMw"
SHEET_NAMES = ["TikTok", "Instagram", "Instagram(Ash)"]
KEY_FILE = "service-account-key.json"

def clean_name(name):
    """
    Very aggressive cleaning for human-like names:
    - Strips concatenated brand terms (Alexishopfinds -> Alexis).
    - Discards purely generic/curation accounts.
    - Strips single-letter initials.
    - Stricter consonant-only check.
    """
    if pd.isna(name): return ""
    
    name_str = str(name).strip()
    
    # 1. Remove non-alphanumeric (keep spaces)
    name_str = re.sub(r'[^a-zA-Z\s]', '', name_str)
    name_str = " ".join(name_str.split())
    if not name_str: return ""

    # 2. Discard purely generic curation keywords (Full Name Check)
    # These often indicate an account that isn't a person.
    generic_discard = [
        'dealswith', 'favfinds', 'reviewsof', 'bookreviews', 'internews', 
        'dailyfinds', 'curatedby', 'digitalproducts', 'solutions', 'marketing',
        'production', 'graphic', 'academy', 'design', 'global', 'official',
        'beginner', 'vlogger', 'blogger', 'review', 'world', 'news', 'vlogs',
        'deals', 'finds', 'ideas', 'media', 'digital', 'agency', 'studio',
        'management', 'community', 'creators', 'creator', 'ugc', 'vlog', 
        'vibe', 'lifestyle', 'shop', 'store', 'channel', 'boutique'
    ]
    name_lower = name_str.lower().replace(" ", "")
    for term in generic_discard:
        if term in name_lower:
            # Exception: Keep if it's a common name part (not worth full complexity now)
            return ""

    # 3. Strip standalone generic words and concatenated suffixes
    # This handles "Abbey UGC" -> "Abbey" AND "Alexishopfinds" -> "Alexi"
    generic_words = [
        'ugc', 'creator', 'influencer', 'vlogger', 'blogger', 'tips', 
        'management', 'official', 'studio', 'media', 'agency', 'marketing', 
        'photography', 'production', 'design', 'global', 'digital', 
        'lifestyle', 'daily', 'world', 'vlog', 'creative', 'shop', 'store', 
        'beginner', 'brand', 'channel', 'uk', 'us', 'usa', 'ca'
    ]
    suffixes_to_strip = [
        'shopfinds', 'favfinds', 'finds', 'deals', 'dealswith', 'reviews', 'ugc',
        'official', 'vlogs', 'vlog', 'daily', 'world'
    ] + generic_words
    
    parts = name_str.split()
    new_parts = []
    for word in parts:
        clean_word = word
        # Is the whole word a generic tag? (Step 3a)
        if clean_word.lower() in generic_words:
            continue
            
        # Or does it have a concatenated suffix? (Step 3b)
        changed = True
        while changed:
            changed = False
            for suffix in suffixes_to_strip:
                lower_word = clean_word.lower()
                suffix_len = len(suffix)
                if lower_word.endswith(suffix) and len(lower_word) > suffix_len + 2:
                    clean_word = clean_word[:-suffix_len].strip()
                    changed = True
                    break
        if clean_word:
            new_parts.append(clean_word)
    name_str = " ".join(new_parts)

    # 4. Handle Single-Letter Logic (Afnan C -> Afnan, D H V J -> Dhvj)
    parts = name_str.split()
    originally_joined = False
    if len(parts) > 1 and all(len(p) == 1 for p in parts):
        name_str = "".join(parts)
        parts = [name_str]
        originally_joined = True
    elif len(parts) > 1:
        # Remove single letter initials unless it's a joinable set
        parts = [p for p in parts if len(p) > 1]
        name_str = " ".join(parts)

    # 5. Red Flag Blacklist for remaining words
    red_flags = [
        'Shop', 'Store', 'Mall', 'Outlet', 'Brand', 'Academy', 'School', 
        'University', 'Portal', 'Channel', 'Productions', 'Marketing', 'Digital'
    ]
    if len(parts) >= 2:
        for flag in red_flags:
            if re.search(rf'\b{flag}\b', name_str, re.IGNORECASE):
                return ""

    # 6. Final Word Selection & Stricter Vowel Check
    name_str = " ".join(name_str.split())
    final_parts = []
    for word in name_str.split():
        word_len = len(word)
        # Handle detection: Words > 3 chars should have vowels (including Y)
        if word_len > 3 and not re.search(r'[aeiouyAEIOUY]', word):
            if not originally_joined:
                continue
        # Also limit to only First Last
        final_parts.append(word)
        if len(final_parts) >= 2:
            break
            
    name_str = " ".join(final_parts)
    
    # 7. Final check (min 3 chars total)
    if len(name_str) < 3:
        return ""
        
    return name_str.title()
        
    return name_str.title()

def find_column(columns, patterns):
    """
    Finds a column name in a list of columns that matches any of the patterns.
    Prioritizes exact matches, then case-insensitive matches, then partial matches.
    """
    # 1. Look for Exact Match first
    for pattern in patterns:
        for col in columns:
            if str(pattern) == str(col):
                return col
                
    # 2. Look for Case-Insensitive Exact Match
    for pattern in patterns:
        for col in columns:
            if str(pattern).lower() == str(col).lower():
                return col
                
    # 3. Look for Partial Match (but avoid common overlaps like 'name' in 'username')
    for pattern in patterns:
        pattern_lower = str(pattern).lower()
        for col in columns:
            col_lower = str(col).lower()
            if pattern_lower == "name" and "username" in col_lower:
                continue # Skip username if we want Name
            if pattern_lower in col_lower:
                return col
    return None

def fetch_from_sheets():
    """
    Fetches data from Google Sheets using the provided credentials and spreadsheet ID.
    """
    if not os.path.exists(KEY_FILE):
        print(f"Error: Credentials file '{KEY_FILE}' not found.")
        print("Please place the service account JSON key in the 'Daily Email' folder.")
        return []

    print(f"Connecting to Google Sheets (ID: {SPREADSHEET_ID})...")
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        credentials = Credentials.from_service_account_file(KEY_FILE, scopes=scopes)
        gc = gspread.authorize(credentials)
        sh = gc.open_by_key(SPREADSHEET_ID)
        
        dataframes = []
        for sheet_name in SHEET_NAMES:
            print(f"  Fetching sheet: {sheet_name}...")
            try:
                worksheet = sh.worksheet(sheet_name)
                data = worksheet.get_all_records()
                if not data:
                    print(f"    Warning: Sheet '{sheet_name}' is empty.")
                    continue
                
                df = pd.DataFrame(data)
                
                # Robustly find "Name" and "Email" columns
                name_col = find_column(df.columns, ["Name", "Full Name", "Username", "Profile Name"])
                email_col = find_column(df.columns, ["Email", "E-mail", "Contact"])
                
                if name_col and email_col:
                    temp_df = df[[name_col, email_col]].copy()
                    temp_df.columns = ["Name", "Email"] # Unify column names
                    # Normalize platform name
                    platform_display = "Instagram" if "Instagram" in sheet_name else sheet_name
                    temp_df.insert(0, "Platform", platform_display) # Add Platform at the beginning
                    dataframes.append(temp_df)
                else:
                    print(f"    Warning: Required columns not found in '{sheet_name}'.")
                    print(f"    Available: {list(df.columns)}")
            except gspread.exceptions.WorksheetNotFound:
                print(f"    Error: Worksheet '{sheet_name}' not found.")
            except Exception as e:
                print(f"    Error processing '{sheet_name}': {e}")
        
        return dataframes
    except Exception as e:
        print(f"Error connecting to Google Sheets: {e}")
        return []

def process_influencer_data():
    """
    Main function to process influencer data.
    """
    print("--- Influencer Data Processor (Cloud Sync) ---")

    # 1. Fetch data from Google Sheets
    dataframes = fetch_from_sheets()

    if not dataframes:
        print("\nError: No valid data found to process.")
        print("Check your sheet names, column names, and service account permissions.")
        return

    # 2. Merge all datasets
    print("\nMerging datasets...")
    merged_df = pd.concat(dataframes, ignore_index=True)

    # 3. Filter and Clean
    print("Cleaning data...")
    
    # Remove rows where Email is empty
    merged_df = merged_df.dropna(subset=["Email"])
    merged_df = merged_df[merged_df["Email"].astype(str).str.strip() != ""]

    # First, keep a copy of the original names for reporting ignored ones
    merged_df["Original Name"] = merged_df["Name"]
    merged_df["Name"] = merged_df["Name"].apply(clean_name)
    
    # Identify ignored rows: Name is empty after cleaning
    ignored_df = merged_df[merged_df["Name"].astype(str).str.strip() == ""].copy()
    ignored_df = ignored_df[["Platform", "Original Name", "Email"]]
    
    # Filter out rows where Name is empty after cleaning for the main dataset
    merged_df = merged_df[merged_df["Name"].astype(str).str.strip() != ""]
    
    # Remove duplicate emails from main dataset
    count_before = len(merged_df)
    merged_df = merged_df.drop_duplicates(subset=["Email"], keep="first")
    count_after = len(merged_df)
    removed_count = count_before - count_after
    if removed_count > 0:
        print(f"  Removed {removed_count} duplicate records (duplicates or cross-platform).")

    # 4. Sort results alphabetically by Name
    print("Sorting results...")
    merged_df = merged_df.sort_values(by="Name")

    # 5. Save output
    OUTPUT_FILE = "final_creator_contacts.xlsx"
    
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d")
    HISTORY_DIR = "history"
    if not os.path.exists(HISTORY_DIR):
        os.makedirs(HISTORY_DIR)

    # Final cleanup of columns for main file
    final_df = merged_df[["Platform", "Name", "Email"]]
    
    # Save main results
    print(f"Saving results to {OUTPUT_FILE} with professional styling...")
    save_with_styling(final_df, OUTPUT_FILE)
    
    # Save ignored results
    IGNORED_FILE = "ignored_contacts.xlsx"
    print(f"Saving ignored contacts to {IGNORED_FILE}...")
    save_ignored_data(ignored_df, IGNORED_FILE)
    
    # Archive a copy with date
    archive_path = os.path.join(HISTORY_DIR, f"contacts_{timestamp}.xlsx")
    save_with_styling(final_df, archive_path)

    # Final summary
    print(f"\nSUCCESS! Processed {len(final_df)} unique human names.")
    print(f"Main file: {os.path.abspath(OUTPUT_FILE)}")
    print(f"Ignored file: {os.path.abspath(IGNORED_FILE)}")
    print(f"Archive copy: {os.path.abspath(archive_path)}")

def save_with_styling(df, path):
    """Saves a dataframe to Excel with professional styling."""
    if df.empty:
        # Create empty file if no data
        df.to_excel(path, index=False)
        return

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        
        # Access the openpyxl workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        # Style definitions
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12, name='Calibri')
        data_font = Font(size=11, name='Calibri')
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply styling to headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Apply styling to data rows and auto-adjust column widths
        for i, col in enumerate(df.columns):
            # Convert index to excel column letter (A, B, C...)
            column_letter = ""
            temp_i = i
            while temp_i >= 0:
                column_letter = chr(65 + (temp_i % 26)) + column_letter
                temp_i = (temp_i // 26) - 1
            
            max_length = len(str(col)) + 4 
            
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=i+1)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def save_ignored_data(df, path):
    """Special styling for ignored contacts."""
    save_with_styling(df, path)

if __name__ == "__main__":
    process_influencer_data()
