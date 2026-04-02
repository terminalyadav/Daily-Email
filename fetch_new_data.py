import pandas as pd
import os
import gspread
from google.oauth2.service_account import Credentials
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configuration (from process_influencers.py)
SPREADSHEET_ID = "1Q8Pw_a88RRZZ9dpaREMAgf5b0RqRqSkQdaJsLGWivMw"
SHEET_NAMES = ["TikTok", "Instagram", "Instagram(Ash)", "Tik-Tok(Ash)"]
KEY_FILE = "service-account-key.json"
OUTPUT_FILE = "new_contacts.xlsx"

def get_existing_emails():
    """Reads emails from existing main and ignored files to ensure we only fetch new ones."""
    emails = set()
    files = ["final_creator_contacts.xlsx", "ignored_contacts.xlsx"]
    for f in files:
        if os.path.exists(f):
            try:
                df = pd.read_excel(f)
                if "Email" in df.columns:
                    emails.update(df["Email"].dropna().astype(str).str.strip().tolist())
            except Exception as e:
                print(f"Error reading {f}: {e}")
    return emails

def find_column(columns, patterns):
    """Finds a column name based on common patterns."""
    for pattern in patterns:
        for col in columns:
            if str(pattern).lower() == str(col).lower():
                return col
    for pattern in patterns:
        pattern_lower = str(pattern).lower()
        for col in columns:
            if pattern_lower in str(col).lower():
                return col
    return None

def save_with_styling(df, path):
    """Saves a dataframe to Excel with professional styling."""
    if df.empty:
        df.to_excel(path, index=False)
        return

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='New Data')
        workbook = writer.book
        worksheet = writer.sheets['New Data']
        
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12, name='Calibri')
        data_font = Font(size=11, name='Calibri')
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for i, col in enumerate(df.columns):
            column_letter = chr(65 + i) # Simple A, B, C...
            max_length = max(len(str(col)) + 4, 15)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=i+1)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def fetch_new_data():
    if not os.path.exists(KEY_FILE):
        print(f"Error: Credentials file '{KEY_FILE}' not found.")
        return

    print("--- Fetching New Influencer Data ---")
    
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        credentials = Credentials.from_service_account_file(KEY_FILE, scopes=scopes)
        gc = gspread.authorize(credentials)
        sh = gc.open_by_key(SPREADSHEET_ID)
        
        existing_emails = get_existing_emails()
        print(f"Loaded {len(existing_emails)} existing unique emails.")

        all_new_data = []
        seen_emails_this_run = set()

        for sheet_name in SHEET_NAMES:
            print(f"  Checking sheet: {sheet_name}...")
            try:
                worksheet = sh.worksheet(sheet_name)
                data = worksheet.get_all_records()
                if not data: continue
                
                df = pd.DataFrame(data)
                
                # Robustly find columns - Updated priority: Username first
                user_col = find_column(df.columns, ["Username", "Profile Name", "Name", "Full Name"])
                email_col = find_column(df.columns, ["Email", "E-mail", "Contact"])
                
                if not email_col or not user_col:
                    print(f"    Warning: Required columns not found in '{sheet_name}'. Skipping.")
                    continue

                for _, row in df.iterrows():
                    email = str(row[email_col]).strip() if pd.notna(row[email_col]) else ""
                    username = str(row[user_col]).strip() if pd.notna(row[user_col]) else ""
                    
                    if email and email not in existing_emails and email not in seen_emails_this_run:
                        if "Instagram" in sheet_name:
                            platform_display = "Instagram"
                        elif "TikTok" in sheet_name or "Tik-Tok" in sheet_name:
                            platform_display = "TikTok"
                        else:
                            platform_display = sheet_name
                        
                        all_new_data.append({
                            "Platform": platform_display,
                            "Username": username,
                            "Email": email
                        })
                        seen_emails_this_run.add(email)
            except Exception as e:
                print(f"    Error processing '{sheet_name}': {e}")
        
        if all_new_data:
            new_df = pd.DataFrame(all_new_data)
            print(f"\nFound {len(new_df)} new entries. Saving to {OUTPUT_FILE}...")
            save_with_styling(new_df, OUTPUT_FILE)
            print("SUCCESS!")
        else:
            print("\nNo new influencer data found.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    fetch_new_data()
