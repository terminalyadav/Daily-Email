"""
FastAPI Backend — Daily Email Dashboard
Endpoints:
  GET  /health
  GET  /api/dates            → all available dates with counts
  GET  /api/emails/{date}    → records for a specific date (YYYY-MM-DD)
  GET  /api/stats            → overall totals + per-platform
  GET  /api/download/{date}  → .xlsx file download for a date
"""
import io
import json
import os
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

import pandas as pd
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

# ─── Config ──────────────────────────────────────────────────────────────────
GITHUB_TOKEN = os.getenv("GITHUB_PAT", "")
GITHUB_REPO  = os.getenv("GITHUB_REPO", "terminalyadav/Daily-Email")

# ─── App setup ───────────────────────────────────────────────────────────────
app = FastAPI(title="Daily Email Dashboard API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # tighten in prod: replace with your Vercel URL
    allow_methods=["*"],
    allow_headers=["*"],
)

# data/ folder now resides inside backend/
DATA_DIR = Path(__file__).parent / "data"


# ─── Helpers ─────────────────────────────────────────────────────────────────

def load_json(date_str: str) -> dict:
    path = DATA_DIR / f"{date_str}.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"No data found for {date_str}")
    with open(path) as f:
        return json.load(f)


def list_all_dates() -> list[dict]:
    """Return all date JSON files sorted newest-first."""
    if not DATA_DIR.exists():
        return []
    result = []
    for f in sorted(DATA_DIR.glob("*.json"), reverse=True):
        try:
            with open(f) as fh:
                d = json.load(fh)
            result.append({
                "date":             d.get("date", f.stem),
                "total":            d.get("total", 0),
                "tiktok_count":     d.get("tiktok_count", 0),
                "instagram_count":  d.get("instagram_count", 0),
                "generated_at":     d.get("generated_at", ""),
            })
        except Exception:
            pass
    return result


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat() + "Z"}


@app.get("/api/dates")
def get_dates():
    """List all available dates with summary counts."""
    dates = list_all_dates()
    return {
        "count":    len(dates),
        "dates":    dates,
    }


@app.get("/api/emails/{date_str}")
def get_emails(date_str: str):
    """
    Get all email records for a specific date.
    date_str format: YYYY-MM-DD
    """
    # validate format
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Date must be in YYYY-MM-DD format")
    return load_json(date_str)


@app.get("/api/stats")
def get_stats():
    """Overall statistics across all dates."""
    dates = list_all_dates()
    total_all = sum(d["total"] for d in dates)
    tiktok    = sum(d["tiktok_count"]    for d in dates)
    instagram = sum(d["instagram_count"] for d in dates)
    return {
        "total_records":   total_all,
        "tiktok_total":    tiktok,
        "instagram_total": instagram,
        "days_recorded":   len(dates),
        "latest_date":     dates[0]["date"] if dates else None,
    }


@app.get("/api/download/{date_str}")
def download_xlsx(date_str: str):
    """Download a .xlsx file for a specific date."""
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Date must be in YYYY-MM-DD format")

    data = load_json(date_str)
    records = data.get("records", [])

    if not records:
        raise HTTPException(status_code=404, detail="No records for this date")

    df = pd.DataFrame(records, columns=["platform", "username", "email"])
    df.columns = ["Platform", "Username", "Email"]

    # Style the Excel file
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.sheets["Data"]
        hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
        data_font = Font(size=11, name="Calibri")
        align     = Alignment(horizontal="left", vertical="center")
        border    = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = len(col_name) + 4
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.font = data_font
                cell.alignment = align
                cell.border = border
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    output.seek(0)
    filename = f"{date_str} emails.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/trigger")
def trigger_github_workflow():
    """Trigger the GitHub Action manually to fetch fresh latest emails."""
    if not GITHUB_TOKEN:
        raise HTTPException(status_code=500, detail="Backend configuration missing: GITHUB_PAT is not set.")
    
    url = f"https://api.github.com/repos/{GITHUB_REPO}/actions/workflows/daily_report.yml/dispatches"
    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json"
    }
    data = json.dumps({"ref": "main"}).encode("utf-8")
    
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        urllib.request.urlopen(req)
        return {"status": "ok", "message": "Triggered successfully."}
    except urllib.error.HTTPError as e:
        err_msg = e.read().decode("utf-8")
        raise HTTPException(status_code=e.code, detail=f"Failed to trigger workflow: {err_msg}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

